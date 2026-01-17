from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _create_styles():
    """Create reusable style definitions."""
    thin_border = Side(style="thin", color="000000")
    return {
        "title_font": Font(bold=True, size=14),
        "section_font": Font(bold=True, size=11),
        "header_font": Font(bold=True, size=10),
        "header_fill": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        "warning_fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "subtotal_font": Font(bold=True),
        "border": Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border),
        "center_align": Alignment(horizontal="center", vertical="center"),
        "right_align": Alignment(horizontal="right", vertical="center"),
        "left_align": Alignment(horizontal="left", vertical="center"),
    }


def _apply_header_row(ws, row: int, columns: List[str], styles: dict):
    """Apply formatting to a header row."""
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.border = styles["border"]
        cell.alignment = styles["center_align"]


def _apply_data_row(ws, row: int, values: List[Any], styles: dict, alignments: List[str] = None):
    """Apply formatting to a data row."""
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = styles["border"]
        if alignments and col_idx <= len(alignments):
            align_type = alignments[col_idx - 1]
            cell.alignment = styles.get(f"{align_type}_align", styles["left_align"])


def _set_column_widths(ws, widths: List[int]):
    """Set column widths."""
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _format_number(value: float, decimals: int = 2) -> str:
    """Format number with specified decimal places."""
    if value is None:
        return "-"
    return f"{value:,.{decimals}f}"


def _format_currency(value: float) -> str:
    """Format currency value."""
    if value is None:
        return "-"
    return f"{value:,.2f}"


def _format_percentage(value: float) -> str:
    """Format percentage value."""
    if value is None:
        return "-"
    return f"{value:.1f}%"


def build_mrp_excel(mrp_data: Dict[str, Any]) -> BytesIO:
    """Generate formatted Excel report from structured MRP data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MRP Raporu"
    styles = _create_styles()

    header = mrp_data.get("header", {})
    summary = mrp_data.get("summary", {})
    lines = mrp_data.get("lines", [])
    bom_summary = mrp_data.get("bom_summary", {})

    current_row = 1

    # === SECTION 1: REPORT HEADER ===
    ws.cell(row=current_row, column=1, value="MRP RAPORU").font = styles["title_font"]
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    current_row += 2

    # Header info
    header_info = [
        ("Proje:", header.get("project_name", "-")),
        ("İş Emri No:", f"#{header.get('work_order_id', '-')}"),
        ("Rapor Tarihi:", header.get("generated_at", "-")[:10] if header.get("generated_at") else "-"),
        ("Satır Sayısı:", header.get("line_count", 0)),
        ("Toplam Miktar:", header.get("total_quantity", 0)),
    ]
    for label, value in header_info:
        ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=value)
        current_row += 1

    current_row += 1

    # === SECTION 2: EXECUTIVE SUMMARY ===
    ws.cell(row=current_row, column=1, value="ÖZET BİLGİLER").font = styles["section_font"]
    current_row += 1

    material = summary.get("material", {})
    cost = summary.get("cost", {})

    summary_data = [
        ("Toplam Sac Alanı:", f"{_format_number(material.get('sheet_area_m2', 0), 3)} m²"),
        ("Toplam Sac Ağırlığı:", f"{_format_number(material.get('sheet_mass_kg', 0), 3)} kg"),
        ("Toplam Yalıtım Alanı:", f"{_format_number(material.get('insulation_area_m2', 0), 3)} m²"),
        ("Tahmini BOM Maliyeti:", f"{_format_currency(cost.get('bom_total', 0))} TL"),
        ("Maliyet Tamlığı:", _format_percentage(bom_summary.get("metrics", {}).get("cost_completeness_pct", 0))),
    ]
    for label, value in summary_data:
        ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=value)
        current_row += 1

    if not cost.get("cost_complete", True):
        current_row += 1
        warning_cell = ws.cell(row=current_row, column=1, value="⚠ Bazı malzemelerin maliyeti eksik - tahmin tam değil")
        warning_cell.fill = styles["warning_fill"]
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)

    current_row += 2

    # === SECTION 3: LINE DETAILS ===
    ws.cell(row=current_row, column=1, value="SATIŞ KALEMLERİ DETAYI").font = styles["section_font"]
    current_row += 1

    line_columns = ["#", "Ürün", "Miktar", "Sac Alanı (m²)", "Sac Ağırlığı (kg)", "Yalıtım Alanı (m²)"]
    _apply_header_row(ws, current_row, line_columns, styles)
    current_row += 1

    line_alignments = ["center", "left", "center", "right", "right", "right"]
    for line in lines:
        totals = line.get("totals", {})
        row_values = [
            line.get("line_number", "-"),
            line.get("product_name", "-"),
            line.get("quantity", 0),
            _format_number(totals.get("sheet_area_m2", 0), 3),
            _format_number(totals.get("sheet_mass_kg", 0), 3),
            _format_number(totals.get("insulation_area_m2", 0), 3),
        ]
        _apply_data_row(ws, current_row, row_values, styles, line_alignments)
        current_row += 1

    # Line subtotals
    subtotal_values = [
        "TOPLAM",
        "",
        sum(l.get("quantity", 0) for l in lines),
        _format_number(material.get("sheet_area_m2", 0), 3),
        _format_number(material.get("sheet_mass_kg", 0), 3),
        _format_number(material.get("insulation_area_m2", 0), 3),
    ]
    _apply_data_row(ws, current_row, subtotal_values, styles, line_alignments)
    for col in range(1, 7):
        ws.cell(row=current_row, column=col).font = styles["subtotal_font"]
    current_row += 2

    # === SECTION 4: BOM SUMMARY - PRICED ITEMS ===
    priced_items = bom_summary.get("priced_items", [])
    if priced_items:
        ws.cell(row=current_row, column=1, value="MALZEME LİSTESİ (FİYATLI)").font = styles["section_font"]
        current_row += 1

        bom_columns = ["Malzeme", "Birim", "Miktar", "Birim Fiyat", "Toplam", "Pay (%)"]
        _apply_header_row(ws, current_row, bom_columns, styles)
        current_row += 1

        bom_alignments = ["left", "center", "right", "right", "right", "right"]
        for item in priced_items:
            row_values = [
                item.get("name", "-"),
                item.get("unit", "-"),
                _format_number(item.get("total_quantity", 0), 2),
                _format_currency(item.get("cost_per_unit", 0)),
                _format_currency(item.get("total_cost", 0)),
                _format_percentage(item.get("cost_share_pct", 0)),
            ]
            _apply_data_row(ws, current_row, row_values, styles, bom_alignments)
            current_row += 1

        # BOM subtotal
        subtotal_values = [
            "TOPLAM",
            "",
            "",
            "",
            _format_currency(cost.get("bom_total", 0)),
            "100%",
        ]
        _apply_data_row(ws, current_row, subtotal_values, styles, bom_alignments)
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).font = styles["subtotal_font"]
        current_row += 2

    # === SECTION 5: BOM SUMMARY - UNPRICED ITEMS ===
    unpriced_items = bom_summary.get("unpriced_items", [])
    if unpriced_items:
        ws.cell(row=current_row, column=1, value="MALZEME LİSTESİ (FİYAT EKSİK)").font = styles["section_font"]
        current_row += 1

        unpriced_columns = ["Malzeme", "Birim", "Miktar", "Durum"]
        _apply_header_row(ws, current_row, unpriced_columns, styles)
        current_row += 1

        unpriced_alignments = ["left", "center", "right", "center"]
        for item in unpriced_items:
            row_values = [
                item.get("name", "-"),
                item.get("unit", "-"),
                _format_number(item.get("total_quantity", 0), 2),
                "Fiyat Gerekli",
            ]
            _apply_data_row(ws, current_row, row_values, styles, unpriced_alignments)
            # Apply warning fill to unpriced rows
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).fill = styles["warning_fill"]
            current_row += 1

        current_row += 1

    # Set column widths for better readability
    _set_column_widths(ws, [25, 15, 15, 18, 18, 12])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
